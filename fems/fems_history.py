
# 20250115 fems history read and fetch used energy and produced energy to calculate self-sufficiency

import asyncio
import base64
import json
import openpyxl
import os
import re
import uuid
import warnings

from datetime import datetime, timedelta
from dotenv import load_dotenv
from websockets.asyncio.client import connect

auth_uuid: uuid = uuid.uuid4()
auth: str = '''{{"jsonrpc":"2.0","id":"{}","method":"authenticateWithPassword","params":{{"username":"{}","password": "{}"}}}}'''

yesterday_value: datetime = datetime.now() - timedelta(1)
yesterday_json: str = yesterday_value.strftime('%Y-%m-%d')
yesterday_prefix: str = yesterday_value.strftime('%Y%m%d_')
edge_uuid: uuid = uuid.uuid4()
history_uuid: uuid = uuid.uuid4()
history: str = '''{{"jsonrpc":"2.0","id":"{}","method":"edgeRpc","params":{{"edgeId":"0","payload":{{"jsonrpc":"2.0","id":"{}","method":"queryHistoricTimeseriesExportXlxs","params":{{"timezone":-3600,"fromDate":"{}","toDate":"{}"}}}}}}}}'''

debug: bool = False
error: str
user: str
password: str
websocket_uri: str
spreadsheet_path: str
spreadsheet_filename: str
result_filename: str


def yet() -> str:
  return datetime.now().strftime("%Y%m%d %H%M%S.%f")


async def calculate():
    async with connect(websocket_uri) as websocket:
        request=auth.format(auth_uuid, "x", "user")
        if debug: print(f"{yet()} >>> {request}")
        await websocket.send(request)
        response = await websocket.recv()
        if debug: print(f"{yet()} <<< {response}")
        request=history.format(edge_uuid, history_uuid, yesterday_json, yesterday_json)
        if debug: print(f"{yet()} >>> {request}")
        await websocket.send(request)
        response = await websocket.recv()
        if debug: print(f"{yet()} <<< {response}")
        data=json.loads(response)
        base64data = data['result']['payload']['result']['payload']
        if debug: print(f"{yet()} <<< {base64data}")
        xlsx = base64.b64decode(base64data)
        with open(spreadsheet_filename, "wb") as file:
            file.write(xlsx)
        with warnings.catch_warnings():
            warnings.filterwarnings("ignore", category=UserWarning, module=re.escape('openpyxl.styles.stylesheet'))
            wb_obj = openpyxl.load_workbook(spreadsheet_filename)
        sheet_obj = wb_obj.active
        used_energy_obj = sheet_obj.cell(row=7, column=7)
        generated_energy_obj = sheet_obj.cell(row=7, column=4)
        independence = float(generated_energy_obj.value) / float(used_energy_obj.value) * 100
        if debug: print(f"{yet()} <<< {generated_energy_obj.value} {used_energy_obj.value} {independence:.1f}")
        with open(result_filename, "w") as file:
            file.write(f"{generated_energy_obj.value} {used_energy_obj.value} {independence:.1f}")


def load_settings() -> bool:
    load_dotenv()
    global user, password, websocket_uri, spreadsheet_path, spreadsheet_filename, result_filename, error
    user = os.getenv('USER')
    password = os.getenv('PASSWORD')
    websocket_uri = os.getenv('WEBSOCKET_URI')
    spreadsheet_path = os.getenv('SPREADSHEET_PATH')
    prefix = 'Nothing was done. Check your .env file, these are undefined: '
    error = prefix
    if not user:
        error += "USER "
    if not password:
        error += "PASSWORD "
    if not websocket_uri:
        error += "WEBSOCKET_URI "
    if not spreadsheet_path:
        spreadsheet_path=""
        error += "SPREADSHEET_PATH "
    spreadsheet_path = spreadsheet_path + (datetime.now() - timedelta(1)).strftime('%Y%m%d_')
    spreadsheet_filename = spreadsheet_path + 'fems_history1.xlsx'
    result_filename = spreadsheet_path + 'fems_history1.txt'
    return len(error) <= len(prefix)

if __name__ == "__main__":
    if load_settings():
        asyncio.run(calculate())
    else:
        print(error)
